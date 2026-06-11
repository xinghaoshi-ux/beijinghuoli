from datetime import date, datetime
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_user
from app.core.exceptions import AppException, NotFoundError
from app.models.competition_item import CompetitionItem
from app.models.player import Player
from app.models.team import Team
from app.models.upload_batch import UploadBatch
from app.schemas.result import VersusMatchCreate
from app.schemas.upload import ConfirmUploadRequest
from app.services.match_service import MatchService

router = APIRouter(prefix="/api/v1/admin/uploads", tags=["uploads"])


@router.post("", status_code=201)
async def upload_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    if not file.filename:
        raise AppException("文件必填", "UPLOAD_FILE_REQUIRED", 422)
    if not file.filename.endswith(".xlsx"):
        raise AppException("文件类型不支持", "UPLOAD_FILE_TYPE_INVALID", 400)

    upload_dir = Path(settings.UPLOAD_DIR)
    upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{uuid4().hex}.xlsx"
    file_path = upload_dir / stored_name
    file_path.write_bytes(await file.read())

    upload = UploadBatch(filename=file.filename, file_path=str(file_path), status="pending")
    db.add(upload)
    await db.commit()
    await db.refresh(upload)

    await _parse_upload(db, upload)
    return {"data": _upload_data(upload)}


@router.get("/{upload_id}")
async def get_upload(
    upload_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    upload = await _get_upload(db, upload_id)
    return {"data": _upload_data(upload)}


@router.get("/{upload_id}/preview")
async def get_preview(
    upload_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    upload = await _get_upload(db, upload_id)
    if upload.status not in ("parsed", "imported"):
        raise AppException("上传状态不允许当前操作", "UPLOAD_STATUS_INVALID", 409)
    return {"data": upload.preview_data or []}


@router.post("/{upload_id}/confirm")
async def confirm_upload(
    upload_id: int,
    body: ConfirmUploadRequest,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    upload = await _get_upload(db, upload_id)
    if upload.status != "parsed":
        raise AppException("上传状态不允许当前操作", "UPLOAD_STATUS_INVALID", 409)
    if not body.confirmed_rows:
        raise AppException("确认导入行不能为空", "UPLOAD_CONFIRM_EMPTY", 422)

    preview = upload.preview_data or []
    imported = 0
    service = MatchService(db)
    for row in preview:
        if row["row_number"] not in body.confirmed_rows:
            continue
        if row.get("row_status") == "error":
            continue
        request = VersusMatchCreate(
            match_date=row.get("match_date"),
            sequence_no=row.get("sequence_no"),
            court=row.get("court"),
            group_name=row.get("group_name"),
            age_group=row.get("age_group"),
            item_id=row["item_id"],
            team_a_id=row["team_a_id"],
            team_b_id=row["team_b_id"],
            team_a_player_ids=row.get("team_a_player_ids") or [],
            team_b_player_ids=row.get("team_b_player_ids") or [],
            team_a_score=row["team_a_score"],
            team_b_score=row["team_b_score"],
            note=row.get("note"),
        )
        await service.create_match(request, source_type="excel", upload_batch_id=upload.id)
        imported += 1

    upload.status = "imported"
    await db.commit()
    return {
        "data": {
            "upload_id": upload.id,
            "status": upload.status,
            "imported_count": imported,
            "ignored_count": len(body.ignored_rows),
        },
        "message": "ok",
    }


@router.post("/{upload_id}/cancel")
async def cancel_upload(
    upload_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    upload = await _get_upload(db, upload_id)
    if upload.status not in ("pending", "parsed", "failed"):
        raise AppException("上传状态不允许当前操作", "UPLOAD_STATUS_INVALID", 409)
    upload.status = "cancelled"
    await db.commit()
    return {"data": _upload_data(upload), "message": "ok"}


async def _get_upload(db: AsyncSession, upload_id: int) -> UploadBatch:
    upload = (
        await db.execute(select(UploadBatch).where(UploadBatch.id == upload_id))
    ).scalar_one_or_none()
    if not upload:
        raise NotFoundError("上传批次不存在", "UPLOAD_NOT_FOUND")
    return upload


def _upload_data(upload: UploadBatch) -> dict:
    return {
        "id": upload.id,
        "filename": upload.filename,
        "status": upload.status,
        "total_rows": upload.total_rows,
        "valid_rows": upload.valid_rows,
        "error_rows": upload.error_rows,
        "error_log": upload.error_log,
    }


async def _parse_upload(db: AsyncSession, upload: UploadBatch) -> None:
    upload.status = "pending"
    await db.commit()
    try:
        preview = await _parse_excel_rows(db, upload.file_path)
        upload.preview_data = preview
        upload.status = "parsed"
        upload.total_rows = len(preview)
        upload.valid_rows = sum(1 for row in preview if row["row_status"] != "error")
        upload.error_rows = sum(1 for row in preview if row["row_status"] == "error")
    except Exception as exc:
        upload.status = "failed"
        upload.error_log = str(exc)
    await db.commit()
    await db.refresh(upload)


async def _parse_excel_rows(db: AsyncSession, file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    header_row = next(
        ws.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    headers = [_normalize_header(value) for value in header_row]
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(value is None for value in row):
            continue
        parsed = await _parse_one_row(db, idx, row, headers)
        rows.append(parsed)
    wb.close()
    return rows


async def _parse_one_row(
    db: AsyncSession,
    row_number: int,
    row: tuple,
    headers: list[str],
) -> dict:
    by_header = _row_getter(row, headers)
    match_date = _date_value(by_header("比赛日期"))
    sequence_no = (
        _int_value(by_header("场序"))
        if by_header("场序") is not None
        else _int_cell(row, 0)
    )
    court = _text_value(by_header("场地")) or _cell(row, 1)
    group_name = (
        _text_value(by_header("年龄组"))
        or _text_value(by_header("组别"))
        or _cell(row, 2)
    )
    item_name = _text_value(by_header("项目名称")) or _cell(row, 3)
    team_a_name = (
        _text_value(by_header("队伍A"))
        or _text_value(by_header("A队"))
        or _cell(row, 4)
    )
    team_a_player_names = [
        _text_value(by_header(name))
        for name in (
            "队伍A球员1",
            "队伍A球员2",
            "队伍A球员3",
            "队伍A球员4",
            "A队球员1",
            "A队球员2",
            "A队球员3",
            "A队球员4",
        )
    ]
    if not any(team_a_player_names):
        team_a_player_names = [_cell(row, col) for col in range(5, 9)]
    team_a_player_names = [name for name in team_a_player_names if name]
    team_a_score = (
        _int_value(by_header("队伍A得分"))
        if by_header("队伍A得分") is not None
        else _int_value(by_header("A队比分"))
        if by_header("A队比分") is not None
        else _int_cell(row, 9)
    )
    team_b_name = (
        _text_value(by_header("队伍B"))
        or _text_value(by_header("B队"))
        or _cell(row, 10)
    )
    team_b_player_names = [
        _text_value(by_header(name))
        for name in (
            "队伍B球员1",
            "队伍B球员2",
            "队伍B球员3",
            "队伍B球员4",
            "B队球员1",
            "B队球员2",
            "B队球员3",
            "B队球员4",
        )
    ]
    if not any(team_b_player_names):
        team_b_player_names = [_cell(row, col) for col in range(11, 15)]
    team_b_player_names = [name for name in team_b_player_names if name]
    team_b_score = (
        _int_value(by_header("队伍B得分"))
        if by_header("队伍B得分") is not None
        else _int_value(by_header("B队比分"))
        if by_header("B队比分") is not None
        else _int_cell(row, 15)
    )
    note = _text_value(by_header("备注")) or _cell(row, 16)

    errors = []
    if match_date is None:
        errors.append("比赛日期必填或格式无效")
    item = None
    if item_name:
        item = (
            await db.execute(select(CompetitionItem).where(CompetitionItem.name == item_name))
        ).scalar_one_or_none()
    if not item:
        errors.append("项目不存在")
    if team_a_score is None or team_a_score < 0 or team_b_score is None or team_b_score < 0:
        errors.append("比分无效")
    if team_a_score is not None and team_b_score is not None and team_a_score == team_b_score:
        errors.append("比分不能为平局")

    team_a_id = None
    team_b_id = None
    if not team_a_name:
        errors.append("A 队名称必填")
    else:
        team_a = (
            await db.execute(select(Team).where(Team.name == team_a_name))
        ).scalar_one_or_none()
        if not team_a:
            errors.append(f"A 队不存在：{team_a_name}")
        else:
            team_a_id = team_a.id
    if not team_b_name:
        errors.append("B 队名称必填")
    else:
        team_b = (
            await db.execute(select(Team).where(Team.name == team_b_name))
        ).scalar_one_or_none()
        if not team_b:
            errors.append(f"B 队不存在：{team_b_name}")
        else:
            team_b_id = team_b.id
    if team_a_id and team_b_id and team_a_id == team_b_id:
        errors.append("对阵球队不能相同")

    expected_count = item.player_count if item else None
    if expected_count and len(team_a_player_names) != expected_count:
        errors.append("A 队球员人数与项目人数不匹配")
    if expected_count and len(team_b_player_names) != expected_count:
        errors.append("B 队球员人数与项目人数不匹配")
    team_a_player_ids = await _lookup_players(db, team_a_player_names, errors, "A 队")
    team_b_player_ids = await _lookup_players(db, team_b_player_names, errors, "B 队")
    all_player_ids = team_a_player_ids + team_b_player_ids
    if len(all_player_ids) != len(set(all_player_ids)):
        errors.append("同一场比赛球员不能重复")

    return {
        "row_number": row_number,
        "match_date": match_date.isoformat() if match_date else None,
        "sequence_no": sequence_no,
        "court": court,
        "group_name": group_name,
        "age_group": group_name,
        "item_name": item_name,
        "item_id": item.id if item else None,
        "team_a_name": team_a_name,
        "team_a_id": team_a_id,
        "team_a_player_names": team_a_player_names,
        "team_a_player_ids": team_a_player_ids,
        "team_a_score": team_a_score,
        "team_b_name": team_b_name,
        "team_b_id": team_b_id,
        "team_b_player_names": team_b_player_names,
        "team_b_player_ids": team_b_player_ids,
        "team_b_score": team_b_score,
        "score_text": (
            f"{team_a_score}:{team_b_score}"
            if team_a_score is not None and team_b_score is not None
            else None
        ),
        "note": note,
        "row_status": "error" if errors else "normal",
        "error_message": "；".join(errors) if errors else None,
    }


async def _lookup_players(
    db: AsyncSession, names: list[str], errors: list[str], label: str
) -> list[int]:
    player_ids = []
    for name in names:
        player = (await db.execute(select(Player).where(Player.name == name))).scalar_one_or_none()
        if not player:
            errors.append(f"{label}球员不存在：{name}")
        else:
            player_ids.append(player.id)
    return player_ids


def _cell(row: tuple, index: int) -> str | None:
    if len(row) <= index or row[index] is None:
        return None
    value = str(row[index]).strip()
    return value or None


def _int_cell(row: tuple, index: int) -> int | None:
    raw = _cell(row, index)
    if raw is None:
        return None
    try:
        return int(raw)
    except ValueError:
        return None


def _normalize_header(value) -> str:
    return str(value).strip().replace(" ", "") if value is not None else ""


def _row_getter(row: tuple, headers: list[str]):
    header_index = {header: index for index, header in enumerate(headers) if header}

    def get(name: str):
        index = header_index.get(_normalize_header(name))
        if index is None or len(row) <= index:
            return None
        return row[index]

    return get


def _text_value(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _int_value(value) -> int | None:
    text = _text_value(value)
    if text is None:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _date_value(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text_value(value)
    if text is None:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
